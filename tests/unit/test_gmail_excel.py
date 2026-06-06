# tests/unit/test_gmail_excel.py
"""Unit tests for Excel report generation."""
import openpyxl
import pytest
from openacm.plugins.gmail_classifier.excel_export import generate_excel

SAMPLE_STATS = {
    "period": {"from": "2026-06-01", "to": "2026-06-30", "total_emails": 10},
    "volume_by_day": [
        {"date": "2026-06-01", "count": 4},
        {"date": "2026-06-02", "count": 6},
    ],
    "by_category": [
        {"id": 1, "name": "Trabajo", "color": "#f00",
         "total": 7, "read": 6, "replied": 3, "ai_classified": 7},
        {"id": 2, "name": "Spam", "color": "#0f0",
         "total": 3, "read": 3, "replied": 0, "ai_classified": 3},
    ],
    "top_senders": [
        {"email": "boss@co.com", "name": "Jefe", "count": 5},
        {"email": "other@co.com", "name": "Otro", "count": 3},
    ],
    "reply_rate": {"total": 10, "replied": 3, "rate": 0.3},
    "autoreply": {
        "suggestions_generated": 8,
        "drafts_saved": 2,
        "examples_learned": 3,
        "avg_use_count": 1.5,
    },
}


def test_workbook_has_five_sheets():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    assert set(wb.sheetnames) == {"Resumen", "Volumen diario", "Por categoría", "Top remitentes", "Auto-reply"}


def test_resumen_sheet_contains_total_emails():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    ws = wb["Resumen"]
    values = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert 10 in values  # total_emails


def test_volumen_diario_has_correct_rows():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    ws = wb["Volumen diario"]
    assert ws.cell(2, 1).value == "2026-06-01"
    assert ws.cell(2, 2).value == 4
    assert ws.cell(3, 1).value == "2026-06-02"
    assert ws.cell(3, 2).value == 6


def test_por_categoria_has_correct_rows():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    ws = wb["Por categoría"]
    assert ws.cell(2, 1).value == "Trabajo"
    assert ws.cell(2, 2).value == 7
    assert ws.cell(3, 1).value == "Spam"


def test_top_remitentes_has_correct_rows():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    ws = wb["Top remitentes"]
    assert ws.cell(2, 1).value == "boss@co.com"
    assert ws.cell(2, 3).value == 5


def test_autoreply_sheet_values():
    buf = generate_excel(SAMPLE_STATS, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(buf)
    ws = wb["Auto-reply"]
    values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    assert values["Sugerencias generadas"] == 8
    assert values["Borradores guardados"] == 2
    assert values["Ejemplos aprendidos"] == 3
