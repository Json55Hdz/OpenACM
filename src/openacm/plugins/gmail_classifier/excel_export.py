# src/openacm/plugins/gmail_classifier/excel_export.py
"""Generate Excel report for Gmail Classifier stats."""
from __future__ import annotations
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _header_style(ws, row: int, cols: int) -> None:
    """Bold + light-gray fill for header row."""
    fill = PatternFill("solid", fgColor="D9D9D9")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def generate_excel(stats: dict[str, Any], from_date: str, to_date: str) -> BytesIO:
    """Return a BytesIO containing the .xlsx report."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Resumen ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 18
    ws1.append([f"Reporte Gmail Classifier — {from_date} → {to_date}"])
    ws1.cell(1, 1).font = Font(bold=True, size=13)
    ws1.append([])
    ws1.append(["Métrica", "Valor"])
    _header_style(ws1, 3, 2)
    rows = [
        ("Total emails recibidos",           stats["period"]["total_emails"]),
        ("Total respondidos",                stats["reply_rate"]["replied"]),
        ("Tasa de respuesta",                f"{stats['reply_rate']['rate'] * 100:.1f}%"),
        ("Sugerencias IA generadas",         stats["autoreply"]["suggestions_generated"]),
        ("Borradores guardados",             stats["autoreply"]["drafts_saved"]),
        ("Ejemplos aprendidos",              stats["autoreply"]["examples_learned"]),
        ("Uso promedio por ejemplo",         stats["autoreply"]["avg_use_count"]),
    ]
    for r in rows:
        ws1.append(list(r))

    # ── Sheet 2: Volumen diario ───────────────────────────────────────────
    ws2 = wb.create_sheet("Volumen diario")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 20
    ws2.append(["Fecha", "Emails recibidos"])
    _header_style(ws2, 1, 2)
    for row in stats["volume_by_day"]:
        ws2.append([row["date"], row["count"]])

    # ── Sheet 3: Por categoría ────────────────────────────────────────────
    ws3 = wb.create_sheet("Por categoría")
    for col, width in zip("ABCDE", [20, 10, 12, 16, 20]):
        ws3.column_dimensions[col].width = width
    ws3.append(["Categoría", "Total", "Leídos", "Respondidos", "Clasificados por IA"])
    _header_style(ws3, 1, 5)
    for cat in stats["by_category"]:
        ws3.append([cat["name"], cat["total"], cat["read"], cat["replied"], cat["ai_classified"]])

    # ── Sheet 4: Top remitentes ───────────────────────────────────────────
    ws4 = wb.create_sheet("Top remitentes")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 18
    ws4.append(["Email", "Nombre", "Emails enviados"])
    _header_style(ws4, 1, 3)
    for s in stats["top_senders"]:
        ws4.append([s["email"], s["name"], s["count"]])

    # ── Sheet 5: Auto-reply ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Auto-reply")
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 18
    ws5.append(["Métrica", "Valor"])
    _header_style(ws5, 1, 2)
    ar = stats["autoreply"]
    ws5.append(["Sugerencias generadas",    ar["suggestions_generated"]])
    ws5.append(["Borradores guardados",      ar["drafts_saved"]])
    ws5.append(["Ejemplos aprendidos",       ar["examples_learned"]])
    ws5.append(["Uso promedio por ejemplo",  ar["avg_use_count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
